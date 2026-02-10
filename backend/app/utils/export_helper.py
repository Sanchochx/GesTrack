"""
Helper para exportación de datos a CSV y Excel

US-INV-003: CA-6 - Export functionality
"""
import csv
import io
from datetime import datetime
from flask import Response


class ExportHelper:
    """Helper para exportar datos a diferentes formatos"""

    @staticmethod
    def export_to_csv(data, columns, filename_prefix='export'):
        """
        Exporta datos a formato CSV

        Args:
            data: Lista de diccionarios con los datos
            columns: Lista de tuplas (key, header_name) que define las columnas
            filename_prefix: Prefijo para el nombre del archivo

        Returns:
            Flask Response con el archivo CSV
        """
        # Crear un buffer de texto en memoria
        si = io.StringIO()

        # Crear writer CSV
        writer = csv.writer(si, quoting=csv.QUOTE_ALL)

        # Escribir encabezados
        headers = [col[1] for col in columns]
        writer.writerow(headers)

        # Escribir datos
        for row in data:
            row_data = []
            for col_key, _ in columns:
                value = row.get(col_key, '')
                # Formatear valores especiales
                if value is None:
                    value = ''
                elif isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, bool):
                    value = 'Sí' if value else 'No'
                row_data.append(str(value))
            writer.writerow(row_data)

        # Preparar respuesta
        output = si.getvalue()
        si.close()

        # Generar nombre de archivo con timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'{filename_prefix}_{timestamp}.csv'

        # Crear respuesta
        response = Response(
            output,
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename={filename}',
                'Content-Type': 'text/csv; charset=utf-8'
            }
        )

        return response

    @staticmethod
    def export_inventory_movements_to_csv(movements):
        """
        Exporta movimientos de inventario a CSV (US-INV-003 CA-6)

        Args:
            movements: Lista de diccionarios de movimientos

        Returns:
            Flask Response con el archivo CSV
        """
        # Definir columnas para exportación
        columns = [
            ('created_at', 'Fecha y Hora'),
            ('product_name', 'Producto'),
            ('product_sku', 'SKU'),
            ('movement_type', 'Tipo de Movimiento'),
            ('quantity', 'Cantidad'),
            ('previous_stock', 'Stock Anterior'),
            ('new_stock', 'Stock Resultante'),
            ('user_name', 'Usuario'),
            ('reason', 'Motivo'),
            ('reference', 'Referencia'),
            ('notes', 'Notas')
        ]

        return ExportHelper.export_to_csv(
            data=movements,
            columns=columns,
            filename_prefix='historial_inventario'
        )

    @staticmethod
    def export_to_excel(data, columns, filename_prefix='export', sheet_name='Datos'):
        """
        Exporta datos a formato Excel (XLSX)

        Requiere openpyxl: pip install openpyxl

        Args:
            data: Lista de diccionarios con los datos
            columns: Lista de tuplas (key, header_name)
            filename_prefix: Prefijo para el nombre del archivo
            sheet_name: Nombre de la hoja

        Returns:
            Flask Response con el archivo Excel
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Si openpyxl no está instalado, retornar CSV como fallback
            return ExportHelper.export_to_csv(data, columns, filename_prefix)

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Estilo para encabezados
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2e7d32', end_color='2e7d32', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Escribir encabezados
        headers = [col[1] for col in columns]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Escribir datos
        for row_num, row in enumerate(data, 2):
            for col_num, (col_key, _) in enumerate(columns, 1):
                value = row.get(col_key, '')

                # Formatear valores especiales
                if value is None:
                    value = ''
                elif isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, bool):
                    value = 'Sí' if value else 'No'

                ws.cell(row=row_num, column=col_num, value=value)

        # Ajustar ancho de columnas
        for col_num in range(1, len(columns) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 15

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generar nombre de archivo con timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'{filename_prefix}_{timestamp}.xlsx'

        # Crear respuesta
        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename={filename}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )

        return response

    @staticmethod
    def export_inventory_movements_to_excel(movements):
        """
        Exporta movimientos de inventario a Excel (US-INV-003 CA-6)

        Args:
            movements: Lista de diccionarios de movimientos

        Returns:
            Flask Response con el archivo Excel
        """
        # Definir columnas para exportación
        columns = [
            ('created_at', 'Fecha y Hora'),
            ('product_name', 'Producto'),
            ('product_sku', 'SKU'),
            ('movement_type', 'Tipo de Movimiento'),
            ('quantity', 'Cantidad'),
            ('previous_stock', 'Stock Anterior'),
            ('new_stock', 'Stock Resultante'),
            ('user_name', 'Usuario'),
            ('reason', 'Motivo'),
            ('reference', 'Referencia'),
            ('notes', 'Notas')
        ]

        return ExportHelper.export_to_excel(
            data=movements,
            columns=columns,
            filename_prefix='historial_inventario',
            sheet_name='Movimientos de Inventario'
        )

    @staticmethod
    def export_inventory_value_report_to_excel(value_data, categories, top_products, evolution):
        """
        US-INV-005 CA-7: Exporta reporte completo de valor del inventario a Excel

        Args:
            value_data: Diccionario con valor total del inventario
            categories: Lista de categorías con su valor
            top_products: Lista de productos de mayor valor
            evolution: Lista de snapshots históricos

        Returns:
            Flask Response con el archivo Excel
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.chart import LineChart, Reference
        except ImportError:
            # Si openpyxl no está instalado, retornar error
            from flask import jsonify
            return jsonify({
                'success': False,
                'error': {
                    'code': 'LIBRARY_NOT_INSTALLED',
                    'message': 'openpyxl no está instalado'
                }
            }), 500

        # Crear workbook
        wb = Workbook()

        # Estilos comunes
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='1976d2', end_color='1976d2', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        title_font = Font(bold=True, size=14, color='1976d2')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # === HOJA 1: Resumen General ===
        ws_summary = wb.active
        ws_summary.title = 'Resumen General'

        # Título
        ws_summary['A1'] = 'REPORTE DE VALOR DEL INVENTARIO'
        ws_summary['A1'].font = Font(bold=True, size=16, color='1976d2')
        ws_summary.merge_cells('A1:D1')

        # Fecha de generación
        ws_summary['A2'] = f'Fecha de generación: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws_summary.merge_cells('A2:D2')

        # Resumen de valor
        ws_summary['A4'] = 'Valor Total del Inventario'
        ws_summary['A4'].font = title_font
        ws_summary['B4'] = value_data.get('formatted_value', '$0.00')
        ws_summary['B4'].font = Font(bold=True, size=14, color='2e7d32')

        ws_summary['A5'] = 'Total de Productos'
        ws_summary['B5'] = value_data.get('total_products', 0)

        ws_summary['A6'] = 'Total de Unidades'
        ws_summary['B6'] = value_data.get('total_quantity', 0)

        # Ajustar anchos
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20

        # === HOJA 2: Desglose por Categoría ===
        ws_categories = wb.create_sheet('Por Categoría')

        # Encabezados
        category_headers = ['Categoría', 'Valor Total', 'Productos', 'Unidades', '% del Total']
        for col_num, header in enumerate(category_headers, 1):
            cell = ws_categories.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Datos
        for row_num, category in enumerate(categories, 2):
            ws_categories.cell(row=row_num, column=1, value=category['category_name']).border = border
            ws_categories.cell(row=row_num, column=2, value=category['total_value']).border = border
            ws_categories.cell(row=row_num, column=3, value=category['product_count']).border = border
            ws_categories.cell(row=row_num, column=4, value=category['total_quantity']).border = border
            ws_categories.cell(row=row_num, column=5, value=f"{category['percentage']:.2f}%").border = border

        # Ajustar anchos
        for col_num in range(1, 6):
            ws_categories.column_dimensions[get_column_letter(col_num)].width = 18

        # === HOJA 3: Top Productos ===
        ws_products = wb.create_sheet('Top Productos')

        # Encabezados
        product_headers = ['#', 'SKU', 'Nombre', 'Categoría', 'Costo Unitario', 'Stock', 'Valor Total']
        for col_num, header in enumerate(product_headers, 1):
            cell = ws_products.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Datos
        for row_num, (index, product) in enumerate(enumerate(top_products, 1), 2):
            ws_products.cell(row=row_num, column=1, value=index).border = border
            ws_products.cell(row=row_num, column=2, value=product['sku']).border = border
            ws_products.cell(row=row_num, column=3, value=product['name']).border = border
            ws_products.cell(row=row_num, column=4, value=product['category_name']).border = border
            ws_products.cell(row=row_num, column=5, value=product['cost_price']).border = border
            ws_products.cell(row=row_num, column=6, value=product['stock_quantity']).border = border
            ws_products.cell(row=row_num, column=7, value=product['total_value']).border = border

        # Ajustar anchos
        ws_products.column_dimensions['A'].width = 5
        ws_products.column_dimensions['B'].width = 15
        ws_products.column_dimensions['C'].width = 30
        ws_products.column_dimensions['D'].width = 20
        ws_products.column_dimensions['E'].width = 15
        ws_products.column_dimensions['F'].width = 10
        ws_products.column_dimensions['G'].width = 15

        # === HOJA 4: Evolución Histórica ===
        if evolution and len(evolution) > 0:
            ws_evolution = wb.create_sheet('Evolución Histórica')

            # Encabezados
            evolution_headers = ['Fecha', 'Valor Total', 'Productos', 'Unidades']
            for col_num, header in enumerate(evolution_headers, 1):
                cell = ws_evolution.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Datos
            for row_num, snapshot in enumerate(evolution, 2):
                # Parsear fecha
                snapshot_date = snapshot['snapshot_date']
                if isinstance(snapshot_date, str):
                    try:
                        from dateutil import parser
                        snapshot_date = parser.parse(snapshot_date).strftime('%Y-%m-%d %H:%M')
                    except:
                        pass

                ws_evolution.cell(row=row_num, column=1, value=snapshot_date).border = border
                ws_evolution.cell(row=row_num, column=2, value=snapshot['total_value']).border = border
                ws_evolution.cell(row=row_num, column=3, value=snapshot['total_products']).border = border
                ws_evolution.cell(row=row_num, column=4, value=snapshot['total_quantity']).border = border

            # Ajustar anchos
            for col_num in range(1, 5):
                ws_evolution.column_dimensions[get_column_letter(col_num)].width = 18

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generar nombre de archivo con timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'reporte_valor_inventario_{timestamp}.xlsx'

        # Crear respuesta
        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename={filename}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )

        return response

    @staticmethod
    def export_inventory_data_to_csv(products_data):
        """
        US-INV-009 CA-3: Exporta datos completos del inventario a CSV

        Args:
            products_data: Lista de diccionarios con datos de productos

        Returns:
            Flask Response con el archivo CSV
        """
        columns = [
            ('sku', 'SKU'),
            ('nombre', 'Nombre'),
            ('categoria', 'Categoría'),
            ('stock_actual', 'Stock Actual'),
            ('stock_reservado', 'Stock Reservado'),
            ('stock_disponible', 'Stock Disponible'),
            ('punto_reorden', 'Punto de Reorden'),
            ('precio_costo', 'Precio Costo'),
            ('precio_venta', 'Precio Venta'),
            ('valor_total', 'Valor Total'),
            ('estado', 'Estado'),
            ('ultima_actualizacion', 'Última Actualización'),
            ('proveedor', 'Proveedor Principal')
        ]

        return ExportHelper.export_to_csv(
            data=products_data,
            columns=columns,
            filename_prefix='inventario'
        )

    @staticmethod
    def export_inventory_data_to_excel(products_data, summary_data):
        """
        US-INV-009 CA-6: Exporta datos completos del inventario a Excel con formato enriquecido

        Args:
            products_data: Lista de diccionarios con datos de productos
            summary_data: Diccionario con datos de resumen

        Returns:
            Flask Response con el archivo Excel
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, numbers
            from openpyxl.utils import get_column_letter
        except ImportError:
            return ExportHelper.export_inventory_data_to_csv(products_data)

        wb = Workbook()

        # Estilos
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1976d2', end_color='1976d2', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        currency_format = '#,##0.00'

        # === HOJA 1: Inventario ===
        ws = wb.active
        ws.title = 'Inventario'

        columns = [
            ('sku', 'SKU', 15),
            ('nombre', 'Nombre', 30),
            ('categoria', 'Categoría', 20),
            ('stock_actual', 'Stock Actual', 14),
            ('stock_reservado', 'Stock Reservado', 16),
            ('stock_disponible', 'Stock Disponible', 16),
            ('punto_reorden', 'Punto de Reorden', 16),
            ('precio_costo', 'Precio Costo', 14),
            ('precio_venta', 'Precio Venta', 14),
            ('valor_total', 'Valor Total', 14),
            ('estado', 'Estado', 15),
            ('ultima_actualizacion', 'Última Actualización', 22),
            ('proveedor', 'Proveedor Principal', 20)
        ]

        # Encabezados
        for col_num, (_, header, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Datos
        currency_cols = {8, 9, 10}  # precio_costo, precio_venta, valor_total
        for row_num, product in enumerate(products_data, 2):
            for col_num, (key, _, _) in enumerate(columns, 1):
                value = product.get(key, '')
                if value is None:
                    value = ''
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if col_num in currency_cols and isinstance(value, (int, float)):
                    cell.number_format = currency_format

        # Auto-filtros y fila congelada
        if products_data:
            last_col = get_column_letter(len(columns))
            ws.auto_filter.ref = f'A1:{last_col}{len(products_data) + 1}'
        ws.freeze_panes = 'A2'

        # === HOJA 2: Resumen ===
        ws_summary = wb.create_sheet('Resumen')
        title_font = Font(bold=True, size=14, color='1976d2')

        ws_summary['A1'] = 'RESUMEN DE EXPORTACIÓN DE INVENTARIO'
        ws_summary['A1'].font = Font(bold=True, size=16, color='1976d2')
        ws_summary.merge_cells('A1:C1')

        ws_summary['A3'] = 'Fecha de Exportación'
        ws_summary['A3'].font = Font(bold=True)
        ws_summary['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        ws_summary['A5'] = 'Total de Productos'
        ws_summary['A5'].font = Font(bold=True)
        ws_summary['B5'] = summary_data.get('total_products', 0)

        ws_summary['A6'] = 'Valor Total del Inventario'
        ws_summary['A6'].font = Font(bold=True)
        ws_summary['B6'] = summary_data.get('total_value', 0)
        ws_summary['B6'].number_format = currency_format

        ws_summary['A8'] = 'Productos por Estado'
        ws_summary['A8'].font = title_font

        status_counts = summary_data.get('status_counts', {})
        ws_summary['A9'] = 'En Stock (Normal)'
        ws_summary['B9'] = status_counts.get('normal', 0)
        ws_summary['A10'] = 'Stock Bajo'
        ws_summary['B10'] = status_counts.get('low_stock', 0)
        ws_summary['A11'] = 'Sin Stock'
        ws_summary['B11'] = status_counts.get('out_of_stock', 0)

        ws_summary['A13'] = 'Filtro Aplicado'
        ws_summary['A13'].font = Font(bold=True)
        ws_summary['B13'] = summary_data.get('filter_applied', 'Todos los productos')

        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 25

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'inventario_{timestamp}.xlsx'

        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename={filename}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )

        return response

    @staticmethod
    def export_inventory_value_report_to_pdf(value_data, categories, top_products):
        """
        US-INV-005 CA-7: Exporta reporte de valor del inventario a PDF

        Args:
            value_data: Diccionario con valor total del inventario
            categories: Lista de categorías con su valor
            top_products: Lista de productos de mayor valor

        Returns:
            Flask Response con el archivo PDF
        """
        # Por ahora retornar Excel como fallback
        # La implementación de PDF requeriría instalar reportlab u otra librería
        from flask import jsonify
        return jsonify({
            'success': False,
            'error': {
                'code': 'NOT_IMPLEMENTED',
                'message': 'Exportación a PDF no implementada aún. Use Excel.'
            }
        }), 501
